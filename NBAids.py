from openpyxl import load_workbook
from openpyxl import Workbook


def getIDs():
	wb = load_workbook(filename = 'ESPNplayerIDs.xlsx') #load the workbook  
	ws = wb['Names and IDs'] #grab the worksheet from the workbook
	playerIDs = {} #create a dictionary {playeer name -> player ID}

	#loop through all 464 players that are in the workbook put them with their ID's in the dictionary
	for row in range(1,464):
	    playerCell = "A" + str(row)
	    idCell = "B" + str(row)
	    playerArray = ws[playerCell].value.lower().split("-")
	    player = playerArray[0] + " " + playerArray[1]
	    iD = int(ws[idCell].value)
	    playerIDs[player] = iD
	return playerIDs