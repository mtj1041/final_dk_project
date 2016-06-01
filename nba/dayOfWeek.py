from bs4 import BeautifulSoup
from datetime import date
from operator import itemgetter
import requests
import time
import sys
import datetime
import operator
import numpy as np
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
# from openpyxl.compat import range
from openpyxl.cell import get_column_letter
#from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
#import lxml

wb = load_workbook(filename = 'ESPNplayerIDs.xlsx')
ws = wb['Names and IDs']

#Create 3 arrays to hold the data from column R, C, and F of Rancho compiled.xlsx
player_names = []
playerIDs = []
F_Names = []

#Loop through the columns, and add each entry into their respective array. Row # should match up with array index.
for row in range(1,463):
    for col in ["A", "B"]:
        cell = col + str(row)
        names = ws[cell].value
        if col == "A":
            player_names.append(names)
        elif col == "B":
            playerIDs.append(names)


base_url = 'http://espn.go.com/nba/player/gamelog/_/id/'


def getStatsForMatchUp(playerID):
    page = base_url + str(playerID)
    print(page)
    r = requests.get(page)
    #soup = BeautifulSoup(r.text, 'lxml')
    soup = BeautifulSoup(r.text)
    table = soup.find("table",{"class" : "tablehead"})
    data = [[stat.get_text() for stat in t.find_all("td")] for t in table.find_all("tr")]
    for array in data: 
        if len(array) != 17:
            data.remove(array)


    days = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    points = [0,0,0,0,0,0,0]
    rebs = [0,0,0,0,0,0,0]
    assts = [0,0,0,0,0,0,0]
    stls = [0,0,0,0,0,0,0]
    blks = [0,0,0,0,0,0,0]
    tos = [0,0,0,0,0,0,0]
    stats = [rebs, assts, blks, stls, tos, points]
    count = [0,0,0,0,0,0,0]

    for array in data:
        day = array[0][:3] 
        if day in days:
            ind = days.index(day)
            count[ind] += 1
            statcount = -7
            for stat in stats:
                if statcount != -2:
                    stat[ind] += float(array[statcount])
                else:
                    stat[ind] += float(array[-1])
                statcount += 1

    for i in range(len(count)):
        if count[i] != 0:
            for stat in stats:
                stat[i] = stat[i]/count[i]
        else:
            stat[i] = "NA"
    
    scoringRules = [1.25, 1.5, 2, 2, -.5, 1]
    for i in range(len(scoringRules)):
        stats[i] = [scoringRules[i]*elem for elem in stats[i]]

    fantasypoints = []
    dayPoints = 0
    i = 0
    while i < 7:
        for stat in stats:
            dayPoints += stat[i]
        fantasypoints.append(dayPoints)
        dayPoints = 0
        i += 1

    print(fantasypoints)
    print(count)
# data = {}
# for i in range(len(player_names)):
#     data[player_names[i]] = getStatsForMatchUp(playerIDs[i])
    
# newb = Workbook() #Create a new Excel File 
# dest_filename = 'fantasyPointsDaysOfWeek.xlsx' #name the file 
# ws1 = newb.active #determine which worksheet to edit (active one) 
# ws1.title = 'Data' #Title this worksheet

# #function that takes in a dictionary and returns and array of all of the keys (tuples)


# for row in range(1,len(data)): #Openpyxl indexes the rows/columns to start at 1 (i.e first row = 1 NOT 0)
#     _ = ws1.cell(column= 1, row=row, value= player_names[row-1])
#     _ = ws1.cell(column= 2, row=row, value= data[player_names[row-1]][0])
#     _ = ws1.cell(column= 3, row=row, value= data[player_names[row-1]][1])
#     _ = ws1.cell(column= 4, row=row, value= data[player_names[row-1]][2])
#     _ = ws1.cell(column= 5, row=row, value= data[player_names[row-1]][3])
#     _ = ws1.cell(column= 6, row=row, value= data[player_names[row-1]][4])
#     _ = ws1.cell(column= 7, row=row, value= data[player_names[row-1]][5])
#     _ = ws1.cell(column= 8, row=row, value= data[player_names[row-1]][6])
                                                                                       
                                                                                
# newb.save(filename = dest_filename) #Save the new excel file

